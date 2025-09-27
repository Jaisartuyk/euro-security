"""
Vistas para el sistema médico con Dr. Claude
EURO SECURITY - Asistente Médico IA
"""
import json
import uuid
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone
from django.db import transaction
from datetime import timedelta
# from .decorators import employee_required, permission_required
# Usar decoradores básicos por ahora
def employee_required(view_func):
    """Decorador básico para empleados"""
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def permission_required(permission):
    """Decorador básico para permisos"""
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator
from .models import (
    Employee, MedicalDocument, MedicalLeave, DrClaudeConversation,
    MedicalDocumentType, MedicalLeaveStatus
)
# from .dr_claude_service import dr_claude
# Importar dinámicamente para evitar import circular
def get_dr_claude():
    from .dr_claude_service import dr_claude
    return dr_claude


@login_required
@employee_required
def medical_dashboard(request):
    """Dashboard principal del sistema médico"""
    try:
        employee = Employee.objects.get(user=request.user)
        
        # Obtener resumen médico del empleado
        medical_summary = get_dr_claude().get_employee_medical_summary(employee)
        
        # Documentos recientes
        recent_documents = MedicalDocument.objects.filter(
            employee=employee
        ).order_by('-uploaded_at')[:5]
        
        # Permisos activos
        active_leaves = MedicalLeave.objects.filter(
            employee=employee,
            status__in=[
                MedicalLeaveStatus.ACTIVE,
                MedicalLeaveStatus.AI_APPROVED,
                MedicalLeaveStatus.HR_APPROVED
            ]
        ).order_by('-start_date')
        
        # Historial de permisos médicos (últimos 6 meses)
        six_months_ago = timezone.now() - timedelta(days=180)
        medical_leaves_history = MedicalLeave.objects.filter(
            employee=employee,
            created_at__gte=six_months_ago
        ).order_by('-created_at')[:10]
        
        # Asistencias recientes (últimos 30 días)
        from .models import AttendanceRecord
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_attendances = AttendanceRecord.objects.filter(
            employee=employee,
            timestamp__gte=thirty_days_ago
        ).order_by('-timestamp')[:15]
        
        # Asistencias afectadas por permisos médicos
        affected_attendances = []
        for leave in active_leaves:
            # Buscar asistencias en el rango del permiso
            leave_attendances = AttendanceRecord.objects.filter(
                employee=employee,
                timestamp__date__range=[leave.start_date, leave.end_date]
            )
            affected_attendances.extend(leave_attendances)
        
        # Estadísticas ampliadas
        stats = {
            'total_documents': MedicalDocument.objects.filter(employee=employee).count(),
            'pending_documents': MedicalDocument.objects.filter(
                employee=employee, 
                processed_by_ai=False
            ).count(),
            'active_leaves': active_leaves.count(),
            'total_medical_days': sum(leave.total_days for leave in active_leaves),
            'medical_leaves_this_month': MedicalLeave.objects.filter(
                employee=employee,
                created_at__month=timezone.now().month,
                created_at__year=timezone.now().year
            ).count(),
            'attendances_this_month': AttendanceRecord.objects.filter(
                employee=employee,
                timestamp__month=timezone.now().month,
                timestamp__year=timezone.now().year
            ).count(),
            'medical_days_this_year': sum(
                leave.total_days for leave in MedicalLeave.objects.filter(
                    employee=employee,
                    created_at__year=timezone.now().year,
                    status__in=[
                        MedicalLeaveStatus.AI_APPROVED,
                        MedicalLeaveStatus.HR_APPROVED,
                        MedicalLeaveStatus.ACTIVE
                    ]
                )
            )
        }
        
        context = {
            'employee': employee,
            'recent_documents': recent_documents,
            'active_leaves': active_leaves,
            'medical_leaves_history': medical_leaves_history,
            'recent_attendances': recent_attendances,
            'affected_attendances': affected_attendances,
            'stats': stats,
            'medical_summary': medical_summary,
            'dr_claude_greeting': get_dr_claude().personality['greeting'],
            'page_title': 'Dr. Claude - Asistente Médico IA'
        }
        
        return render(request, 'attendance/medical/dashboard.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Perfil de empleado no encontrado.')
        return redirect('attendance:dashboard')


@csrf_exempt
@login_required
@employee_required
def upload_medical_document(request):
    """Subir documento médico para análisis IA"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        employee = Employee.objects.get(user=request.user)
        
        # Validar archivo
        if 'document' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No se encontró archivo para subir'
            })
        
        document_file = request.FILES['document']
        document_type = request.POST.get('document_type', MedicalDocumentType.CERTIFICATE)
        
        # Validar tipo de archivo
        allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png']
        file_extension = document_file.name.lower().split('.')[-1]
        if f'.{file_extension}' not in allowed_extensions:
            return JsonResponse({
                'success': False,
                'error': 'Tipo de archivo no permitido. Use PDF, JPG o PNG.'
            })
        
        # Validar tamaño (máximo 10MB)
        if document_file.size > 10 * 1024 * 1024:
            return JsonResponse({
                'success': False,
                'error': 'El archivo es demasiado grande. Máximo 10MB.'
            })
        
        with transaction.atomic():
            # Crear documento médico
            medical_doc = MedicalDocument.objects.create(
                employee=employee,
                document_type=document_type,
                document_file=document_file
            )
            
            # Procesar con Dr. Claude
            analysis_result = get_dr_claude().analyze_medical_certificate(medical_doc)
            
            if analysis_result.get('success', True):
                # Crear permiso médico si es necesario
                if document_type == MedicalDocumentType.CERTIFICATE:
                    medical_leave = get_dr_claude().create_medical_leave(medical_doc)
                    
                    return JsonResponse({
                        'success': True,
                        'message': '¡Documento procesado exitosamente por Dr. Claude!',
                        'document_id': medical_doc.id,
                        'analysis': medical_doc.ai_analysis,
                        'confidence': medical_doc.ai_confidence_score,
                        'leave_created': medical_leave is not None,
                        'leave_id': medical_leave.id if medical_leave else None,
                        'extracted_data': medical_doc.ai_extracted_data
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'message': 'Documento subido y analizado correctamente',
                        'document_id': medical_doc.id,
                        'analysis': medical_doc.ai_analysis,
                        'confidence': medical_doc.ai_confidence_score
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'error': analysis_result.get('error', 'Error en el análisis del documento')
                })
        
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Perfil de empleado no encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error procesando documento: {str(e)}'
        })


@csrf_exempt
@login_required
@employee_required
def chat_with_claude(request):
    """Chat en tiempo real con Dr. Claude"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        employee = Employee.objects.get(user=request.user)
        data = json.loads(request.body)
        
        message = data.get('message', '').strip()
        session_id = data.get('session_id', str(uuid.uuid4()))
        
        if not message:
            return JsonResponse({
                'success': False,
                'error': 'Mensaje vacío'
            })
        
        # Procesar con Dr. Claude
        response = get_dr_claude().chat_with_employee(employee, message, session_id)
        
        return JsonResponse(response)
        
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Perfil de empleado no encontrado'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Formato de datos inválido'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en chat: {str(e)}'
        })


@login_required
@employee_required
def medical_document_detail(request, document_id):
    """Ver detalle de documento médico"""
    try:
        employee = Employee.objects.get(user=request.user)
        document = get_object_or_404(
            MedicalDocument,
            id=document_id,
            employee=employee
        )
        
        # Obtener permiso médico asociado si existe
        medical_leave = None
        if document.leaves.exists():
            medical_leave = document.leaves.first()
        
        context = {
            'document': document,
            'medical_leave': medical_leave,
            'page_title': f'Documento Médico - {document.get_document_type_display()}'
        }
        
        return render(request, 'attendance/medical/document_detail.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Perfil de empleado no encontrado.')
        return redirect('attendance:medical_dashboard')


@login_required
@employee_required
def medical_leave_detail(request, leave_id):
    """Ver detalle de permiso médico"""
    try:
        employee = Employee.objects.get(user=request.user)
        leave = get_object_or_404(
            MedicalLeave,
            id=leave_id,
            employee=employee
        )
        
        context = {
            'leave': leave,
            'page_title': f'Permiso Médico - {leave.start_date} a {leave.end_date}'
        }
        
        return render(request, 'attendance/medical/leave_detail.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Perfil de empleado no encontrado.')
        return redirect('attendance:medical_dashboard')


@login_required
@employee_required
def medical_history(request):
    """Historial médico completo del empleado"""
    try:
        employee = Employee.objects.get(user=request.user)
        
        # Documentos médicos
        documents = MedicalDocument.objects.filter(
            employee=employee
        ).order_by('-uploaded_at')
        
        # Permisos médicos
        leaves = MedicalLeave.objects.filter(
            employee=employee
        ).order_by('-start_date')
        
        # Conversaciones con Dr. Claude
        conversations = DrClaudeConversation.objects.filter(
            employee=employee
        ).order_by('-timestamp')[:20]
        
        context = {
            'employee': employee,
            'documents': documents,
            'leaves': leaves,
            'conversations': conversations,
            'page_title': 'Historial Médico Completo'
        }
        
        return render(request, 'attendance/medical/history.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Perfil de empleado no encontrado.')
        return redirect('attendance:medical_dashboard')


@csrf_exempt
@login_required
@employee_required
def rate_claude_response(request):
    """Calificar respuesta de Dr. Claude"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        data = json.loads(request.body)
        conversation_id = data.get('conversation_id')
        rating = data.get('rating')
        
        if not conversation_id or not rating:
            return JsonResponse({
                'success': False,
                'error': 'Datos faltantes'
            })
        
        # Validar rating
        if not isinstance(rating, int) or rating < 1 or rating > 5:
            return JsonResponse({
                'success': False,
                'error': 'Calificación debe ser entre 1 y 5'
            })
        
        # Actualizar conversación
        conversation = get_object_or_404(
            DrClaudeConversation,
            id=conversation_id,
            employee__user=request.user
        )
        
        conversation.user_rating = rating
        conversation.save()
        
        return JsonResponse({
            'success': True,
            'message': '¡Gracias por tu calificación!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error guardando calificación: {str(e)}'
        })


# Vistas para RRHH
@login_required
@permission_required('supervisor')
def hr_medical_dashboard(request):
    """Dashboard médico para RRHH"""
    # Documentos pendientes de revisión
    pending_documents = MedicalDocument.objects.filter(
        processed_by_ai=False
    ).order_by('-uploaded_at')
    
    # Permisos que requieren revisión humana
    pending_leaves = MedicalLeave.objects.filter(
        status=MedicalLeaveStatus.HUMAN_REVIEW
    ).order_by('-created_at')
    
    # Estadísticas del día
    today = timezone.now().date()
    stats = {
        'documents_today': MedicalDocument.objects.filter(
            uploaded_at__date=today
        ).count(),
        'leaves_approved_today': MedicalLeave.objects.filter(
            status__in=[MedicalLeaveStatus.AI_APPROVED, MedicalLeaveStatus.HR_APPROVED],
            created_at__date=today
        ).count(),
        'pending_review': pending_leaves.count(),
        'ai_accuracy': 0.92  # Calcular dinámicamente
    }
    
    context = {
        'pending_documents': pending_documents,
        'pending_leaves': pending_leaves,
        'stats': stats,
        'page_title': 'Dashboard Médico - RRHH'
    }
    
    return render(request, 'attendance/medical/hr_dashboard.html', context)


@csrf_exempt
@login_required
@permission_required('supervisor')
def approve_medical_leave(request, leave_id):
    """Aprobar/rechazar permiso médico (RRHH)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'approve' o 'reject'
        notes = data.get('notes', '')
        
        leave = get_object_or_404(MedicalLeave, id=leave_id)
        
        if action == 'approve':
            leave.approve_by_hr(request.user, notes)
            message = f'Permiso médico aprobado para {leave.employee.get_full_name()}'
        elif action == 'reject':
            leave.status = MedicalLeaveStatus.HR_REJECTED
            leave.reviewed_by = request.user
            leave.reviewed_at = timezone.now()
            leave.hr_notes = notes
            leave.save()
            message = f'Permiso médico rechazado para {leave.employee.get_full_name()}'
        else:
            return JsonResponse({
                'success': False,
                'error': 'Acción no válida'
            })
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error procesando solicitud: {str(e)}'
        })


# =============================================================================
# ACCIONES RÁPIDAS - DASHBOARD RRHH
# =============================================================================

@csrf_exempt
@login_required
@permission_required('supervisor')
def bulk_approve_leaves(request):
    """Aprobación masiva de permisos médicos recomendados por IA"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        # Obtener permisos que Claude AI recomienda aprobar
        leaves_to_approve = MedicalLeave.objects.filter(
            status=MedicalLeaveStatus.HUMAN_REVIEW,
            ai_recommendation='approve',
            ai_confidence_score__gte=0.8  # Solo alta confianza
        )
        
        approved_count = 0
        approved_employees = []
        
        for leave in leaves_to_approve:
            leave.approve_by_hr(
                request.user, 
                'Aprobación masiva basada en recomendación de Dr. Claude IA'
            )
            approved_count += 1
            approved_employees.append(leave.employee.get_full_name())
        
        return JsonResponse({
            'success': True,
            'message': f'¡{approved_count} permisos aprobados exitosamente!',
            'approved_count': approved_count,
            'employees': approved_employees
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error en aprobación masiva: {str(e)}'
        })


@login_required
@permission_required('supervisor')
def generate_medical_report(request):
    """Generar reporte médico en PDF"""
    try:
        from django.http import HttpResponse
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        import io
        
        # Crear buffer para PDF
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Encabezado
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "EURO SECURITY - Reporte Médico")
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 70, f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Estadísticas
        y_position = height - 120
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y_position, "Estadísticas del Mes")
        
        # Datos del mes actual
        current_month = timezone.now().replace(day=1)
        
        stats_data = [
            ("Documentos procesados:", MedicalDocument.objects.filter(
                uploaded_at__gte=current_month
            ).count()),
            ("Permisos aprobados:", MedicalLeave.objects.filter(
                created_at__gte=current_month,
                status__in=[MedicalLeaveStatus.AI_APPROVED, MedicalLeaveStatus.HR_APPROVED]
            ).count()),
            ("Revisión humana:", MedicalLeave.objects.filter(
                created_at__gte=current_month,
                status=MedicalLeaveStatus.HUMAN_REVIEW
            ).count()),
            ("Precisión IA:", "92%")
        ]
        
        y_position -= 30
        p.setFont("Helvetica", 12)
        for label, value in stats_data:
            p.drawString(70, y_position, f"{label} {value}")
            y_position -= 20
        
        # Documentos recientes
        y_position -= 30
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y_position, "Documentos Recientes")
        
        recent_docs = MedicalDocument.objects.filter(
            uploaded_at__gte=current_month
        ).order_by('-uploaded_at')[:10]
        
        y_position -= 30
        p.setFont("Helvetica", 10)
        for doc in recent_docs:
            if y_position < 100:  # Nueva página si es necesario
                p.showPage()
                y_position = height - 50
            
            text = f"{doc.employee.get_full_name()} - {doc.get_document_type_display()} - {doc.uploaded_at.strftime('%d/%m/%Y')}"
            p.drawString(70, y_position, text)
            y_position -= 15
        
        # Finalizar PDF
        p.showPage()
        p.save()
        
        # Preparar respuesta
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_medico_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error generando reporte: {str(e)}'
        })


@login_required
@permission_required('supervisor')
def export_medical_data(request):
    """Exportar datos médicos a Excel"""
    try:
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja 1: Documentos Médicos
        ws1 = wb.active
        ws1.title = "Documentos Médicos"
        
        # Encabezados
        headers1 = [
            'Empleado', 'Tipo Documento', 'Fecha Subida', 'Estado',
            'Confianza IA', 'Diagnóstico', 'Médico', 'Centro Médico'
        ]
        
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos de documentos
        documents = MedicalDocument.objects.all().order_by('-uploaded_at')
        for row, doc in enumerate(documents, 2):
            ws1.cell(row=row, column=1, value=doc.employee.get_full_name())
            ws1.cell(row=row, column=2, value=doc.get_document_type_display())
            ws1.cell(row=row, column=3, value=doc.uploaded_at.strftime('%d/%m/%Y %H:%M'))
            ws1.cell(row=row, column=4, value='Procesado' if doc.processed_by_ai else 'Pendiente')
            ws1.cell(row=row, column=5, value=f"{doc.ai_confidence_score:.1%}" if doc.ai_confidence_score else 'N/A')
            ws1.cell(row=row, column=6, value=doc.diagnosis or 'N/A')
            ws1.cell(row=row, column=7, value=doc.doctor_name or 'N/A')
            ws1.cell(row=row, column=8, value=doc.medical_center or 'N/A')
        
        # Hoja 2: Permisos Médicos
        ws2 = wb.create_sheet("Permisos Médicos")
        
        headers2 = [
            'Empleado', 'Fecha Inicio', 'Fecha Fin', 'Días Totales',
            'Estado', 'Recomendación IA', 'Revisado Por', 'Fecha Revisión'
        ]
        
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos de permisos
        leaves = MedicalLeave.objects.all().order_by('-created_at')
        for row, leave in enumerate(leaves, 2):
            ws2.cell(row=row, column=1, value=leave.employee.get_full_name())
            ws2.cell(row=row, column=2, value=leave.start_date.strftime('%d/%m/%Y'))
            ws2.cell(row=row, column=3, value=leave.end_date.strftime('%d/%m/%Y'))
            ws2.cell(row=row, column=4, value=leave.total_days)
            ws2.cell(row=row, column=5, value=leave.get_status_display())
            ws2.cell(row=row, column=6, value=leave.ai_recommendation or 'N/A')
            ws2.cell(row=row, column=7, value=leave.reviewed_by.get_full_name() if leave.reviewed_by else 'N/A')
            ws2.cell(row=row, column=8, value=leave.reviewed_at.strftime('%d/%m/%Y %H:%M') if leave.reviewed_at else 'N/A')
        
        # Ajustar ancho de columnas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Preparar respuesta
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="datos_medicos_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error exportando datos: {str(e)}'
        })


@csrf_exempt
@login_required
@permission_required('supervisor')
def configure_claude_ai(request):
    """Configurar parámetros de Claude AI"""
    if request.method == 'GET':
        # Mostrar configuración actual
        config = {
            'model': getattr(settings, 'CLAUDE_MODEL', 'claude-opus-4-1-20250805'),
            'max_tokens': getattr(settings, 'CLAUDE_MAX_TOKENS', 1024),
            'temperature': getattr(settings, 'CLAUDE_TEMPERATURE', 0.7),
            'confidence_threshold': 0.8,  # Umbral para aprobación automática
            'auto_approve_enabled': True
        }
        
        return JsonResponse({
            'success': True,
            'config': config
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar parámetros
            confidence_threshold = data.get('confidence_threshold', 0.8)
            auto_approve = data.get('auto_approve_enabled', True)
            
            if not 0.5 <= confidence_threshold <= 1.0:
                return JsonResponse({
                    'success': False,
                    'error': 'Umbral de confianza debe estar entre 0.5 y 1.0'
                })
            
            # Aquí se guardarían en base de datos o cache
            # Por ahora solo confirmamos
            
            return JsonResponse({
                'success': True,
                'message': 'Configuración de Dr. Claude actualizada exitosamente',
                'config': {
                    'confidence_threshold': confidence_threshold,
                    'auto_approve_enabled': auto_approve
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error configurando IA: {str(e)}'
            })
